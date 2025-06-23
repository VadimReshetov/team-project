import math
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, numbers
from openpyxl.utils import get_column_letter


# Целевая функция
def f(x):
    return (x - 1) ** 2  # Изменено на новую целевую функцию


# Метод дихотомии с подсчётом вычислений
def dichotomy_method(a, b, epsilon):
    iterations = []
    eval_count = 0
    while abs(b - a) > epsilon:
        delta = epsilon / 2
        x1 = (a + b - delta) / 2
        x2 = (a + b + delta) / 2
        f1, f2 = f(x1), f(x2)
        eval_count += 2
        if f1 < f2:
            b = x2
        else:
            a = x1
        iterations.append((a, b, b - a, x1, x2, f1, f2))
    return iterations, eval_count


# Метод золотого сечения с подсчётом вычислений
def golden_section_method(a, b, epsilon):
    phi = (math.sqrt(5) - 1) / 2
    x1 = b - phi * (b - a)
    x2 = a + phi * (b - a)
    f1, f2 = f(x1), f(x2)
    eval_count = 2
    iterations = []
    while abs(b - a) > epsilon:
        if f1 < f2:
            b = x2
            x2, f2 = x1, f1
            x1 = b - phi * (b - a)
            f1 = f(x1)
            eval_count += 1
        else:
            a = x1
            x1, f1 = x2, f2
            x2 = a + phi * (b - a)
            f2 = f(x2)
            eval_count += 1
        iterations.append((a, b, b - a, x1, x2, f1, f2))
    return iterations, eval_count


# Форматирование Excel
def format_sheet(ws):
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

            # Заголовки жирным
            if cell.row == 1:
                cell.font = Font(bold=True)

            # Форматирование столбцов
            if cell.column == 1:  # Столбец "i" (первый)
                cell.number_format = numbers.FORMAT_GENERAL  # Общий формат
            elif isinstance(cell.value, (float, int)):  # Остальные числовые столбцы
                cell.number_format = '0.000000'  # 6 знаков после запятой

    # Автоподбор ширины столбцов
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

# Сохранение результатов для одного epsilon
def save_iterations(method_name, iterations, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = method_name
    # Добавляем столбец i и корректируем названия
    ws.append(["i", "x1", "x2", "f(x1)", "f(x2)", "a_i", "b_i", "b_i - a_i"])

    # Заполняем данные с нумерацией итераций
    for idx, iteration in enumerate(iterations, 1):
        a, b, length, x1, x2, f1, f2 = iteration
        ws.append([idx, x1, x2, f1, f2, a, b, length])

    format_sheet(ws)
    wb.save(filename)


    precisions = [10 ** -n for n in range(2, 9)]  # ε = [1e-2, 1e-3, ..., 1e-8]
    results = []

    for eps in precisions:
        _, dich_count = dichotomy_method(a, b, eps)
        _, gold_count = golden_section_method(a, b, eps)
        results.append((f"{eps:.0e}", dich_count, gold_count))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Сравнение методов"
    ws.append(["Точность", "Дихотомия (вычислений)", "Золотое сечение (вычислений)"])

    for row in results:
        ws.append(row)

    format_sheet(ws)
    wb.save("точность сравнения.xlsx")


# Основная программа
if __name__ == "__main__":
    a, b = -2, 20  # Диапазон
    epsilon = 0.1  # Для точности 0.001

    # Для точности 0.001
    dich_iter, _ = dichotomy_method(a, b, epsilon)
    gold_iter, _ = golden_section_method(a, b, epsilon)

    save_iterations("Дихотомия", dich_iter, "дихотомия.xlsx")
    save_iterations("Золотое сечение", gold_iter, "золотое сечение.xlsx")



    print("Готово")
    print("- дихотомия.xlsx\n- золотое сечение.xlsx\n- точность сравнения.xlsx")
